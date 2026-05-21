import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_connected_bom(output_file=r"E:\Excel Agent\Advanced_Linked_BOM.xlsx"):
    try:
        # 1. DATABASE SHEET (Ang pinagkukunan ng presyo)
        parts_db = {
            "Part_ID": ["P-001", "P-002", "P-003", "P-004", "P-005"],
            "Description": ["Processor", "Memory", "Storage", "Power Supply", "Casing"],
            "Standard_Cost": [250, 80, 120, 60, 45],
            "Supplier": ["Intel", "Samsung", "Crucial", "Corsair", "NZXT"]
        }
        
        # 2. BOM SHEET (Ang project list)
        bom_data = {
            "Level": [0, 1, 1, 1, 1, 1],
            "Part_ID": ["SYS-01", "P-001", "P-002", "P-003", "P-004", "P-005"],
            "Qty": [1, 1, 2, 1, 1, 1]
        }

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            pd.DataFrame(parts_db).to_excel(writer, index=False, sheet_name="Parts_Master")
            pd.DataFrame(bom_data).to_excel(writer, index=False, sheet_name="BOM_Project")
            
            workbook = writer.book
            ws_bom = workbook["BOM_Project"]
            ws_db = workbook["Parts_Master"]
            
            # --- CELL REFERENCING & LINKING FORMULAS ---
            
            # Add Headers for Linked Columns
            ws_bom["D1"] = "Unit_Cost (Linked)"
            ws_bom["E1"] = "Extended_Cost"
            ws_bom["F1"] = "Supplier_Ref"
            
            last_row_bom = len(bom_data["Part_ID"]) + 1
            
            for r in range(2, last_row_bom + 1):
                # A. VLOOKUP Logic: Kuhanin ang Cost mula sa Parts_Master sheet
                # Formula: =VLOOKUP(LookupValue, TableRange, ColumnIndex, False)
                ws_bom[f"D{r}"] = f"=VLOOKUP(B{r}, Parts_Master!A:D, 3, FALSE)"
                
                # B. Supplier Reference (Linking to another sheet)
                ws_bom[f"F{r}"] = f"=VLOOKUP(B{r}, Parts_Master!A:D, 4, FALSE)"
                
                # C. Internal Math: Qty * Linked Cost
                ws_bom[f"E{r}"] = f"=C{r}*D{r}"

            # --- SUMMARY DASHBOARD (Cross-Sheet Linking) ---
            ws_dash = workbook.create_sheet("Dashboard")
            ws_dash["A1"] = "PROJECT SUMMARY"
            ws_dash["A1"].font = Font(bold=True, size=14)
            
            ws_dash["A3"] = "Total BOM Cost:"
            # Direct Cell Reference: Tinuturo nito ang specific cell sa BOM_Project sheet
            # Gagamit tayo ng SUM sa BOM_Project at i-lilink dito
            total_cost_formula = f"=SUM(BOM_Project!E2:E{last_row_bom})"
            ws_dash["B3"] = total_cost_formula
            ws_dash["B3"].number_format = '"$"#,##0.00'
            ws_dash["B3"].font = Font(bold=True)

            # --- GUIDE FOR THE USER (Paano mag-link) ---
            ws_dash["A6"] = "HOW TO LINK CELLS (TIPS):"
            ws_dash["A7"] = "1. Cross-Sheet: Type '=' then click the other sheet and the cell."
            ws_dash["A8"] = "2. Formula: =SheetName!CellAddress (e.g., =Parts_Master!C2)"
            ws_dash["A9"] = "3. Absolute Ref: Use '$' to lock a cell (e.g., $C$2) so it won't change when dragged."

            # Formatting
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            for ws in [ws_bom, ws_db]:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                for col in ws.columns:
                    ws.column_dimensions[get_column_letter(col[0].column)].width = 20

        print(f"✅ Success: Linked BOM System created at {output_file}")
    except Exception as e:
        print(f"❌ Error: {str(e)}")

if __name__ == "__main__":
    generate_connected_bom()
